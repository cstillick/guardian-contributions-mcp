"""Build the OK_PrePrimary_and_Continuing_<date>.xlsx deliverable (Book(Sheet1)
layout) from the store's combined figures.

Conventions (workflow §13):
  - columns: Dist., Candidate, Beg. Balance, Raised, Loan, Expended, End Balance, Notes
  - Accounting number format; End Balance is a LIVE formula =C+D+E−F (not a value)
  - top status banner: refresh date, window, extract as-of, "Ending omits
    continuing spend" caveat, and whether anything changed
  - preserve roster order, (D) markers, blank separator rows between districts
  - versioned filename; archive (never overwrite) any file it replaces
  - highlight cells whose End Balance changed vs the prior version (yellow)
"""
from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from ..config import get_settings
from ..db import session_scope
from ..models import Committee
from ..money import to_cents
from ..reporting_calendar import build_calendar
from ..roster import ROSTER_2026, norm_name, resolve_org_id
from .. import service

ACCOUNTING = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADERS = ["Dist.", "Candidate", "Beg. Balance", "Raised", "Loan", "Expended",
           "End Balance", "Notes"]
YELLOW = PatternFill("solid", fgColor="FFFF00")


def _name_index() -> dict[str, list[str]]:
    with session_scope() as s:
        idx: dict[str, list[str]] = {}
        for c in s.scalars(select(Committee)):
            if c.candidate_name:
                idx.setdefault(norm_name(c.candidate_name), []).append(c.org_id)
        return idx


def _cents_or_none(decimal_str: str | None) -> int | None:
    return to_cents(decimal_str) if decimal_str is not None else None


def _row_for(name: str, name_index) -> dict:
    """Resolve a roster name to combined figures (cents) + note."""
    res = resolve_org_id(name, name_index)
    if not res["org_id"]:
        note = ("No committee found" if res["source"] == "no_committee"
                else "Multiple committees — unresolved" if res["multiple"] else "Unresolved")
        return {"beg": None, "raised": None, "loan": None, "exp": None, "note": note}
    combined = service.get_combined(res["org_id"])
    return {
        "beg": _cents_or_none(combined["beginning"]),
        "raised": _cents_or_none(combined["raised"]),
        "loan": _cents_or_none(combined["loans"]),
        "exp": _cents_or_none(combined["expended"]),
        "note": combined["note"],
    }


def _prior_end_balances(prior: Path | None) -> dict[tuple[str, str], int | None]:
    if not prior or not prior.exists():
        return {}
    wb = load_workbook(prior, data_only=True)
    ws = wb.active
    out: dict[tuple[str, str], int | None] = {}
    for r in ws.iter_rows(min_row=1, values_only=True):
        if r and isinstance(r[0], str) and r[0].startswith(("HD-", "SD-")) and len(r) >= 7:
            out[(r[0], str(r[1]))] = to_cents(r[6]) if r[6] is not None else None
    return out


def build_workbook(year: int | None = None, roster: dict[str, list[str]] | None = None,
                   prior: Path | None = None) -> Workbook:
    settings = get_settings()
    year = year or settings.default_cycle_year
    roster = roster or ROSTER_2026
    with session_scope() as s:
        status = service.refresh_status(year)
    cal = build_calendar(year)
    name_index = _name_index()
    prior_ends = _prior_end_balances(prior)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Banner
    asof = (status.get("extract_as_of") or {}).get("max_receipt_date")
    changed = status.get("changed_since_prev")
    banner = (
        f"OK Pre-Primary + Continuing — built {dt.date.today().isoformat()} | "
        f"continuing window {cal.continuing_start} → {cal.continuing_end} | "
        f"extract as-of {asof} | "
        f"{'no change since prior pull' if changed is False else 'updated'} | "
        f"⚠ Ending omits continuing-period spending"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    bcell = ws.cell(row=1, column=1, value=banner)
    bcell.font = Font(bold=True, size=9)
    bcell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 42

    # Header
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True)

    r = 3
    changed_cells = 0
    for district, names in roster.items():
        for name in names:
            fig = _row_for(name, name_index)
            ws.cell(row=r, column=1, value=district)
            ws.cell(row=r, column=2, value=name)
            for col, key in ((3, "beg"), (4, "raised"), (5, "loan"), (6, "exp")):
                cents = fig[key]
                cell = ws.cell(row=r, column=col, value=(cents / 100 if cents is not None else 0))
                cell.number_format = ACCOUNTING
            # End Balance = LIVE formula
            end = ws.cell(row=r, column=7, value=f"=C{r}+D{r}+E{r}-F{r}")
            end.number_format = ACCOUNTING
            ws.cell(row=r, column=8, value=fig["note"])

            # highlight if End Balance changed vs prior
            beg, raised, loan, exp = (fig["beg"] or 0, fig["raised"] or 0,
                                      fig["loan"] or 0, fig["exp"] or 0)
            new_end = beg + raised + loan - exp
            key = (district, name)
            if key in prior_ends and prior_ends[key] != new_end:
                ws.cell(row=r, column=7).fill = YELLOW
                changed_cells += 1
            r += 1
        r += 1  # blank separator row between districts

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = [8, 22, 14, 14, 12, 14, 14, 40][col - 1]

    wb._guardian_changed_cells = changed_cells  # stash for the caller
    return wb


def write_deliverable(year: int | None = None, out_dir: Path | None = None,
                      prior: Path | None = None) -> Path:
    settings = get_settings()
    out_dir = out_dir or settings.output_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"OK_PrePrimary_and_Continuing_{dt.date.today().isoformat()}.xlsx"
    if out.exists():  # Rule 11: archive, never silently overwrite
        archived = out.with_name(out.stem + f"_ARCHIVED_{dt.datetime.now():%H%M%S}.xlsx")
        shutil.move(str(out), str(archived))
    wb = build_workbook(year=year, prior=prior)
    wb.save(out)
    return out


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(description="Build the combined xlsx deliverable.")
    ap.add_argument("--year", type=int, default=None)
    ap.add_argument("--prior", type=Path, default=None, help="prior xlsx to diff/highlight")
    args = ap.parse_args()
    path = write_deliverable(year=args.year, prior=args.prior)
    print(f"wrote {path}")


if __name__ == "__main__":
    main()
